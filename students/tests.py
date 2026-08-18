import io

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook

from accounts.models import User
from classes.models import ClassRoom
from django.core.exceptions import ValidationError
from institutions.models import Institution

from .models import Student
from .services import ExcelImportError, StudentService, generate_strong_password
from .forms import StudentCreationForm


class StudentServiceTests(TestCase):

    def setUp(self):

        self.institution = Institution.objects.create(
            name="Test School",
            short_name="TS",
        )

        self.classroom = ClassRoom.objects.create(
            name="Grade 5",
            institution=self.institution,
        )

    def _make_excel(self, rows, filename="students.xlsx"):

        workbook = Workbook()
        worksheet = workbook.active

        for row in rows:
            worksheet.append(row)

        buffer = io.BytesIO()
        workbook.save(buffer)

        return SimpleUploadedFile(
            filename,
            buffer.getvalue(),
        )

    def _import(self, excel_file, **kwargs):

        return StudentService.import_students(
            excel_file=excel_file,
            institution=self.institution,
            classroom=self.classroom,
            default_password="Test12345!",
            **kwargs,
        )

    def test_import_english_headers(self):

        excel_file = self._make_excel(
            [
                ["First Name", "Last Name", "Student Code"],
                ["Ali", "Ahmed", "S100001"],
                ["Omar", "Hassan", "S100002"],
            ]
        )

        result = self._import(excel_file)

        self.assertEqual(result["created_count"], 2)
        self.assertEqual(result["skipped_count"], 0)
        self.assertEqual(result["errors"], [])
        self.assertEqual(Student.objects.count(), 2)

    def test_import_arabic_headers(self):

        excel_file = self._make_excel(
            [
                ["الاسم الأول", "الاسم الأخير", "رقم الطالب"],
                ["محمد", "علي", "S100003"],
            ]
        )

        result = self._import(excel_file)

        self.assertEqual(result["created_count"], 1)

        student = Student.objects.get(student_code="S100003")
        self.assertEqual(student.user.first_name, "محمد")
        self.assertEqual(student.user.last_name, "علي")
        self.assertTrue(student.user.must_change_password)

    def test_import_arabic_title_row_above_headers(self):

        excel_file = self._make_excel(
            [
                ["مدرسة الاختبار - قائمة الطلاب"],
                [""],
                ["الاسم الأول", "الاسم الأخير", "رقم الطالب"],
                ["حسين", "مهدي", "S100004"],
            ]
        )

        result = self._import(excel_file)

        self.assertEqual(result["created_count"], 1)
        self.assertTrue(
            Student.objects.filter(student_code="S100004").exists()
        )

    def test_import_full_name_column(self):

        excel_file = self._make_excel(
            [
                ["أسماء الطالبات", "رقم الطالب"],
                ["احلام جاسم محمد ناصر", "S100005"],
                ["اسراء ليث إبراهيم هادي", "S100006"],
            ]
        )

        result = self._import(excel_file)

        self.assertEqual(result["created_count"], 2)

        student = Student.objects.get(student_code="S100005")
        self.assertEqual(student.user.first_name, "احلام")
        self.assertEqual(student.user.last_name, "جاسم محمد ناصر")

    def test_import_gender_column(self):

        excel_file = self._make_excel(
            [
                ["First Name", "Last Name", "Student Code", "Gender"],
                ["Ali", "Ahmed", "S100008", "male"],
                ["Sara", "Hassan", "S100009", "أنثى"],
                ["Noor", "Khalid", "S100010", ""],
            ]
        )

        result = self._import(excel_file)

        self.assertEqual(result["created_count"], 3)

        self.assertEqual(
            Student.objects.get(student_code="S100008").gender,
            Student.Gender.MALE,
        )
        self.assertEqual(
            Student.objects.get(student_code="S100009").gender,
            Student.Gender.FEMALE,
        )

    def test_import_default_gender_fallback(self):

        excel_file = self._make_excel(
            [
                ["First Name", "Last Name", "Student Code"],
                ["Ali", "Ahmed", "S100011"],
            ]
        )

        result = self._import(
            excel_file,
            default_gender=Student.Gender.FEMALE,
        )

        self.assertEqual(result["created_count"], 1)

        self.assertEqual(
            Student.objects.get(student_code="S100011").gender,
            Student.Gender.FEMALE,
        )

    def test_import_skips_duplicate_names(self):

        excel_file = self._make_excel(
            [
                ["First Name", "Last Name", "Student Code"],
                ["Ali", "Ahmed", "S100001"],
            ]
        )

        first = self._import(excel_file)
        second = self._import(excel_file)

        self.assertEqual(first["created_count"], 1)
        self.assertEqual(second["created_count"], 0)
        self.assertEqual(second["skipped_count"], 1)
        self.assertEqual(Student.objects.count(), 1)

    def test_import_skips_duplicates_within_file(self):

        excel_file = self._make_excel(
            [
                ["First Name", "Last Name", "Student Code"],
                ["Ali", "Ahmed", "S100001"],
                ["Ali", "Ahmed", "S100002"],
            ]
        )

        result = self._import(excel_file)

        self.assertEqual(result["created_count"], 1)
        self.assertEqual(result["skipped_count"], 1)
        self.assertEqual(Student.objects.count(), 1)

    def test_import_empty_rows_skipped(self):

        excel_file = self._make_excel(
            [
                ["First Name", "Last Name", "Student Code"],
                [None, None, None],
                [None, None, None],
                ["Huda", "Yasin", "S100007"],
            ]
        )

        result = self._import(excel_file)

        self.assertEqual(result["created_count"], 1)
        self.assertEqual(result["skipped_count"], 2)

    def test_import_rejects_non_excel_file(self):

        excel_file = SimpleUploadedFile(
            "students.csv",
            b"First Name,Last Name\nAli,Ahmed\n",
        )

        with self.assertRaises(ExcelImportError):
            self._import(excel_file)

    def test_import_missing_required_column(self):

        excel_file = self._make_excel(
            [
                ["unknown1", "unknown2"],
                ["a", "b"],
            ]
        )

        with self.assertRaises(ExcelImportError):
            self._import(excel_file)

    def test_import_date_formats(self):

        excel_file = self._make_excel(
            [
                ["First Name", "Last Name", "Student Code", "Date of Birth"],
                ["Ali", "Ahmed", "S100008", "2010-05-01"],
                ["Omar", "Hassan", "S100009", "13/02/2011"],
            ]
        )

        result = self._import(excel_file)

        self.assertEqual(result["created_count"], 2)

        student = Student.objects.get(student_code="S100008")
        self.assertEqual(str(student.date_of_birth), "2010-05-01")

    def test_generate_strong_password(self):

        for _ in range(20):
            password = generate_strong_password()
            self.assertEqual(len(password), 12)
            self.assertTrue(any(c.islower() for c in password))
            self.assertTrue(any(c.isupper() for c in password))
            self.assertTrue(any(c.isdigit() for c in password))
            self.assertTrue(any(c in "!@#$%&*?+=" for c in password))

    def test_create_student_auto_generates_password(self):

        student = StudentService.create_student(
            username="autopass",
            password="",
            first_name="Ali",
            last_name="Ahmed",
            institution=self.institution,
            classroom=self.classroom,
            student_code="C000005",
            gender="male",
            date_of_birth=None,
            address="",
            parent_phone="",
        )

        self.assertTrue(student.user.must_change_password)
        self.assertIsNotNone(student.generated_password)
        self.assertTrue(student.user.check_password(student.generated_password))

    def test_import_auto_generates_passwords(self):

        excel_file = self._make_excel(
            [
                ["First Name", "Last Name", "Student Code"],
                ["Ali", "Ahmed", "S100100"],
                ["Omar", "Hassan", "S100101"],
            ]
        )

        result = StudentService.import_students(
            excel_file=excel_file,
            institution=self.institution,
            classroom=self.classroom,
            default_password="",
        )

        self.assertEqual(result["created_count"], 2)
        self.assertEqual(len(result["credentials"]), 2)

        first = result["credentials"][0]
        self.assertTrue(first["code"].startswith("STU"))
        self.assertEqual(len(first["password"]), 12)

        ali = Student.objects.get(student_code="S100100")
        self.assertTrue(ali.user.check_password(first["password"]))

    def test_create_student_duplicate_username(self):

        StudentService.create_student(
            username="duplicate",
            password="pass12345",
            first_name="Ali",
            last_name="Ahmed",
            institution=self.institution,
            classroom=self.classroom,
            student_code="C000001",
            gender="male",
            date_of_birth=None,
            address="",
            parent_phone="",
        )

        with self.assertRaises(ValidationError):

            StudentService.create_student(
                username="duplicate",
                password="pass12345",
                first_name="Omar",
                last_name="Hassan",
                institution=self.institution,
                classroom=self.classroom,
                student_code="C000002",
                gender="male",
                date_of_birth=None,
                address="",
                parent_phone="",
            )

        self.assertEqual(Student.objects.count(), 1)

    def test_create_student_duplicate_code(self):

        StudentService.create_student(
            username="code_one",
            password="pass12345",
            first_name="Ali",
            last_name="Ahmed",
            institution=self.institution,
            classroom=self.classroom,
            student_code="C000010",
            gender="male",
            date_of_birth=None,
            address="",
            parent_phone="",
        )

        with self.assertRaises(ValidationError):

            StudentService.create_student(
                username="code_two",
                password="pass12345",
                first_name="Omar",
                last_name="Hassan",
                institution=self.institution,
                classroom=self.classroom,
                student_code="C000010",
                gender="male",
                date_of_birth=None,
                address="",
                parent_phone="",
            )


class StudentImportViewTests(TestCase):

    def setUp(self):

        self.admin = User.objects.create_superuser(
            username="admin",
            password="adminpass123",
            first_name="Admin",
        )

        self.institution = Institution.objects.create(
            name="Test School",
            short_name="TS",
        )

        self.classroom = ClassRoom.objects.create(
            name="Grade 5",
            institution=self.institution,
        )

    def _make_excel(self, rows):

        workbook = Workbook()
        worksheet = workbook.active

        for row in rows:
            worksheet.append(row)

        buffer = io.BytesIO()
        workbook.save(buffer)

        return SimpleUploadedFile(
            "students.xlsx",
            buffer.getvalue(),
        )

    def test_import_view_requires_manager(self):

        student_user = User.objects.create_user(
            username="student",
            password="pass12345",
            role=User.Role.STUDENT,
        )
        self.client.force_login(student_user)

        response = self.client.get(
            reverse("students:import"),
        )

        self.assertEqual(response.status_code, 403)

    def test_import_view_success(self):

        self.client.force_login(self.admin)

        excel_file = self._make_excel(
            [
                ["الاسم الأول", "الاسم الأخير", "رقم الطالب"],
                ["محمد", "علي", "V100001"],
                ["احمد", "كريم", "V100002"],
            ]
        )

        response = self.client.post(
            reverse("students:import"),
            {
                "excel_file": excel_file,
                "institution": self.institution.pk,
                "classroom": self.classroom.pk,
                "default_password": "Test12345!",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Student.objects.count(), 2)
        self.assertContains(
            response,
            "تم استيراد 2 طالباً بنجاح.",
        )


class StudentCreationFormTests(TestCase):

    def setUp(self):

        self.institution = Institution.objects.create(
            name="Test School",
            short_name="TS",
        )

        self.classroom = ClassRoom.objects.create(
            name="Grade 5",
            institution=self.institution,
        )

    def _valid_data(self):

        return {
            "first_name": "Ali",
            "last_name": "Ahmed",
            "username": "ali.ahmed",
            "student_code": "S100050",
            "gender": Student.Gender.MALE,
            "password": "Strongpass1",
            "confirm_password": "Strongpass1",
            "institution": self.institution.pk,
            "classroom": self.classroom.pk,
            "date_of_birth": "2010-01-01",
            "address": "",
            "parent_phone": "",
        }

    def test_create_password_optional(self):

        data = self._valid_data()
        data.pop("password")
        data.pop("confirm_password")

        form = StudentCreationForm(
            data,
            institution=self.institution,
        )

        self.assertTrue(form.is_valid())

    def test_create_rejects_mismatched_passwords(self):

        data = self._valid_data()
        data["confirm_password"] = "Different1"

        form = StudentCreationForm(
            data,
            institution=self.institution,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("__all__", form.errors)

    def test_create_valid_form(self):

        form = StudentCreationForm(
            self._valid_data(),
            institution=self.institution,
        )

        self.assertTrue(form.is_valid())

    def test_update_does_not_require_password(self):

        data = self._valid_data()
        data.pop("password")
        data.pop("confirm_password")

        form = StudentCreationForm(
            data,
            is_update=True,
            institution=self.institution,
        )

        self.assertTrue(form.is_valid())


class StudentCredentialsViewTests(TestCase):

    def setUp(self):

        self.admin = User.objects.create_superuser(
            username="admin",
            password="adminpass123",
            first_name="Admin",
        )

        self.institution = Institution.objects.create(
            name="Test School",
            short_name="TS",
        )

        self.classroom = ClassRoom.objects.create(
            name="Grade 5",
            institution=self.institution,
        )

    def _login_manager(self):

        self.client.force_login(self.admin)

    def test_created_requires_manager(self):

        student_user = User.objects.create_user(
            username="student",
            password="pass12345",
            role=User.Role.STUDENT,
        )
        self.client.force_login(student_user)

        response = self.client.get(
            reverse("students:created"),
        )

        self.assertEqual(response.status_code, 403)

    def test_created_without_credentials_redirects(self):

        self._login_manager()

        response = self.client.get(
            reverse("students:created"),
        )

        self.assertEqual(
            response.status_code,
            302,
        )
        self.assertRedirects(
            response,
            reverse("students:list"),
        )

    def test_create_with_auto_password_redirects_to_created(self):

        self._login_manager()

        response = self.client.post(
            reverse("students:create"),
            {
                "first_name": "Ali",
                "last_name": "Ahmed",
                "username": "ali.created",
                "student_code": "V200001",
                "gender": Student.Gender.MALE,
                "institution": self.institution.pk,
                "classroom": self.classroom.pk,
                "date_of_birth": "2010-01-01",
            },
        )

        self.assertRedirects(
            response,
            reverse("students:created"),
        )

        student = Student.objects.get(student_code="V200001")

        credentials = self.client.session[
            "student_credentials"
        ]["credentials"]

        self.assertEqual(len(credentials), 1)
        self.assertEqual(
            credentials[0]["name"],
            "Ali Ahmed",
        )
        self.assertTrue(
            student.user.check_password(
                credentials[0]["password"],
            ),
        )

    def test_created_page_shows_credentials(self):

        self._login_manager()

        session = self.client.session
        session["student_credentials"] = {
            "institution": self.institution.name,
            "at": "2026-01-01T10:00:00+00:00",
            "credentials": [
                {
                    "code": "STU000123",
                    "username": "ali",
                    "name": "Ali Ahmed",
                    "password": "Xk#9mQ2pL!zT",
                }
            ],
        }
        session.save()

        response = self.client.get(
            reverse("students:created"),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "Ali Ahmed",
        )
        self.assertContains(
            response,
            "STU000123",
        )
        self.assertContains(
            response,
            reverse("students:credentials_pdf"),
        )

    def test_credentials_pdf_downloads(self):

        self._login_manager()

        session = self.client.session
        session["student_credentials"] = {
            "institution": self.institution.name,
            "at": "2026-01-01T10:00:00+00:00",
            "credentials": [
                {
                    "code": "STU000123",
                    "username": "ali",
                    "name": "Ali Ahmed",
                    "password": "Xk#9mQ2pL!zT",
                },
                {
                    "code": "STU000124",
                    "username": "sara",
                    "name": "سارة خالد",
                    "password": "vB5@pQn8wR$2",
                },
            ],
        }
        session.save()

        response = self.client.get(
            reverse("students:credentials_pdf"),
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/pdf",
        )
        self.assertIn(
            "attachment",
            response["Content-Disposition"],
        )
        self.assertTrue(
            response.content.startswith(b"%PDF"),
        )
        self.assertIn(
            "student_credentials",
            self.client.session,
        )

        second = self.client.get(
            reverse("students:credentials_pdf"),
        )
        self.assertEqual(second.status_code, 200)
        self.assertTrue(
            second.content.startswith(b"%PDF"),
        )

    def test_credentials_pdf_without_data_redirects(self):

        self._login_manager()

        response = self.client.get(
            reverse("students:credentials_pdf"),
        )

        self.assertEqual(
            response.status_code,
            302,
        )
        self.assertRedirects(
            response,
            reverse("students:list"),
        )

    def test_import_stores_credentials_for_pdf(self):

        self._login_manager()

        workbook = Workbook()
        worksheet = workbook.active

        worksheet.append(
            ["الاسم الأول", "الاسم الأخير", "رقم الطالب"],
        )
        worksheet.append(
            ["محمد", "علي", "V300001"],
        )
        worksheet.append(
            ["احمد", "كريم", "V300002"],
        )

        buffer = io.BytesIO()
        workbook.save(buffer)

        excel_file = SimpleUploadedFile(
            "students.xlsx",
            buffer.getvalue(),
        )

        response = self.client.post(
            reverse("students:import"),
            {
                "excel_file": excel_file,
                "institution": self.institution.pk,
                "classroom": self.classroom.pk,
                "default_password": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Student.objects.count(), 2)

        credentials = self.client.session[
            "student_credentials"
        ]["credentials"]

        self.assertEqual(len(credentials), 2)
        self.assertEqual(len(credentials[0]["password"]), 12)

        pdf_response = self.client.get(
            reverse("students:credentials_pdf"),
        )

        self.assertEqual(pdf_response.status_code, 200)
        self.assertTrue(
            pdf_response.content.startswith(b"%PDF"),
        )
