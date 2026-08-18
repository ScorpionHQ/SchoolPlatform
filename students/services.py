import itertools
import random
import secrets
import string
import unicodedata
from datetime import datetime

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.translation import gettext as _

from openpyxl import load_workbook

from accounts.models import User
from notifications.services import NotificationService
from .models import Student


def generate_strong_password(length=12):

    lowercase = string.ascii_lowercase
    uppercase = string.ascii_uppercase
    digits = string.digits
    symbols = "!@#$%&*?+="

    password = [
        secrets.choice(lowercase),
        secrets.choice(uppercase),
        secrets.choice(digits),
        secrets.choice(symbols),
    ]

    pool = lowercase + uppercase + digits + symbols

    password.extend(
        secrets.choice(pool)
        for _ in range(length - len(password))
    )

    random.SystemRandom().shuffle(password)

    return "".join(password)


class ExcelImportError(ValidationError):
    pass


class StudentService:

    @staticmethod
    @transaction.atomic
    def create_student(
        *,
        username,
        password,
        first_name,
        last_name,
        institution,
        classroom,
        student_code,
        gender,
        date_of_birth,
        address,
        parent_phone,
    ):

        if User.objects.filter(username=username).exists():
            raise ValidationError(_("Username already exists."))

        if Student.objects.filter(student_code=student_code).exists():
            raise ValidationError(_("Student code already exists."))

        generated_password = None

        if not password:
            password = generate_strong_password()
            generated_password = password

        user = User.objects.create_user(
            username=username,
            password=password,
            first_name=first_name,
            last_name=last_name,
            role=User.Role.STUDENT,
            gender=gender,
            must_change_password=generated_password is not None,
        )

        student = Student.objects.create(
            user=user,
            institution=institution,
            classroom=classroom,
            student_code=student_code,
            gender=gender,
            date_of_birth=date_of_birth,
            address=address,
            parent_phone=parent_phone,
        )

        if generated_password:
            student.generated_password = generated_password

        full_name = student.user.get_full_name() or student.student_code

        NotificationService.notify_institution_managers(
            institution,
            ntype=NotificationService.Type.STUDENT,
            title=_("New student added"),
            message=_(
                "Student {name} ({code}) has been added to {institution}."
            ).format(
                name=full_name,
                code=student.student_code,
                institution=institution.name,
            ),
            link=f"/students/{student.pk}/update/",
        )

        NotificationService.notify_user(
            student.user,
            ntype=NotificationService.Type.STUDENT,
            title=_("Welcome!"),
            message=_(
                "Your student account has been created at {institution}. "
                "Please log in and change your password."
            ).format(
                institution=institution.name,
            ),
            link="/students/profile/",
        )

        return student

    @staticmethod
    @transaction.atomic
    def update_student(
        *,
        student,
        username,
        password,
        first_name,
        last_name,
        institution,
        classroom,
        student_code,
        gender,
        date_of_birth,
        address,
        parent_phone,
    ):

        if User.objects.exclude(pk=student.user.pk).filter(
            username=username
        ).exists():
            raise ValidationError(_("Username already exists."))

        if Student.objects.exclude(pk=student.pk).filter(
            student_code=student_code
        ).exists():
            raise ValidationError(_("Student code already exists."))

        student.user.username = username
        student.user.first_name = first_name
        student.user.last_name = last_name
        student.user.gender = gender

        if password:
            student.user.set_password(password)

        student.user.save()

        student.institution = institution
        student.classroom = classroom
        student.student_code = student_code
        student.gender = gender
        student.date_of_birth = date_of_birth
        student.address = address
        student.parent_phone = parent_phone

        student.save()

        full_name = student.user.get_full_name() or student.student_code

        NotificationService.notify_institution_managers(
            institution,
            ntype=NotificationService.Type.STUDENT,
            title=_("Student updated"),
            message=_(
                "The data of student {name} ({code}) has been updated."
            ).format(
                name=full_name,
                code=student.student_code,
            ),
            link=f"/students/{student.pk}/update/",
        )

        NotificationService.student_and_parents(
            student,
            ntype=NotificationService.Type.STUDENT,
            title=_("Profile updated"),
            message=_("Your student data has been updated."),
            student_link="/students/profile/",
            parent_link="/",
        )

        return student

    @staticmethod
    def _normalize_header(value):

        if value is None:
            return ""

        text = str(value).strip().lower()

        text = text.replace("\ufeff", "")
        text = text.replace("\u200f", "")
        text = text.replace("\xa0", " ")

        text = text.replace("أ", "ا")
        text = text.replace("إ", "ا")
        text = text.replace("آ", "ا")
        text = text.replace("ى", "ي")
        text = text.replace("ة", "ه")
        text = text.replace("ئ", "ي")
        text = text.replace("ؤ", "و")
        text = text.replace("ء", "")

        text = "".join(
            char
            for char in text
            if unicodedata.category(char) != "Mn"
        )

        return " ".join(text.split())

    @staticmethod
    def _build_header_map(header_row):

        aliases = {
            "first_name": (
                "first name",
                "firstname",
                "f name",
                "الاسم الاول",
                "الاسم الأول",
                "اسم الطالب",
                "اسم التلميذ",
                "اسم التلميذه",
                "الاسم",
                "المسمي",
                "المسمى",
            ),
            "last_name": (
                "last name",
                "lastname",
                "l name",
                "surname",
                "family name",
                "الاسم الاخير",
                "الاسم الأخير",
                "اسم العائلة",
                "العائلة",
                "اللقب",
                "لقب الطالب",
            ),
            "full_name": (
                "full name",
                "fullname",
                "name",
                "student name",
                "الاسم الكامل",
                "الاسم كامل",
                "اسم الطالب الكامل",
                "اسما الطلاب",
                "اسما الطالبات",
                "اسامي الطلاب",
                "اسامي الطالبات",
                "اسم الطلاب",
                "اسم الطالبات",
                "اسم الطالبه",
                "اسم الطالبة",
                "اسم التلميذ",
                "اسم التلميذة",
                "اسامي التلاميذ",
                "اسامى التلاميذ",
                "أسماء الطلاب",
                "أسماء الطالبات",
                "أسماء التلاميذ",
                "قائمة الاسماء",
                "قائمة الأسماء",
                "قائمة الاسامي",
                "اسامى",
                "أسماء",
                "الاسماء",
                "الأسامي",
                "المسمي الكامل",
                "المسمى الكامل",
            ),
            "student_code": (
                "student code",
                "student id",
                "code",
                "رقم الطالب",
                "رقم التلميذ",
                "كود الطالب",
                "كود التلميذ",
                "رمز الطالب",
                "رقم القيد",
                "الرقم",
                "رقم_student",
            ),
            "gender": (
                "gender",
                "sex",
                "النوع",
                "الجنس",
                "جنس الطالب",
                "نوع الطالب",
            ),
            "username": (
                "username",
                "user name",
                "اسم المستخدم",
                "user_name",
            ),
            "date_of_birth": (
                "date of birth",
                "dob",
                "birth date",
                "birthday",
                "تاريخ الميلاد",
                "تاريخ الولاده",
                "تاريخ الولادة",
                "ميلاد",
                "تاريخ",
            ),
            "parent_phone": (
                "parent phone",
                "phone",
                "mobile",
                "phone number",
                "رقم ولي الامر",
                "رقم ولي الأمر",
                "هاتف ولي الامر",
                "هاتف ولي الأمر",
                "رقم ولي امر",
                "هاتف ولي امر",
                "الموبايل",
                "الهاتف",
                "رقم الجوال",
                "جوال",
                "هاتف",
                "رقم هاتف",
                "رقم الهاتف",
            ),
            "address": (
                "address",
                "عنوان",
                "العنوان",
                "عنوان الطالب",
                "العنوان المنزلي",
            ),
        }

        header_map = {}

        for index, raw_header in enumerate(header_row):

            normalized = StudentService._normalize_header(
                raw_header,
            )

            if not normalized:
                continue

            for field_name, aliases_list in aliases.items():

                if normalized in aliases_list:

                    header_map[field_name] = index

                    break

        return header_map

    @staticmethod
    def _detect_header_row(rows):

        peeked = []

        for _ in range(10):

            try:
                peeked.append(next(rows))
            except StopIteration:
                break

        if not peeked:
            return None, [], 0, []

        best_index = -1
        best_score = -1

        for index, row in enumerate(peeked):

            score = len(
                StudentService._build_header_map(row)
            )

            if score > best_score:

                best_score = score
                best_index = index

        first_content_index = -1

        for index, row in enumerate(peeked):

            if any(
                cell is not None
                and str(cell).strip()
                for cell in row
            ):

                first_content_index = index
                break

        if best_score <= 0:
            return None, [], first_content_index, peeked

        header_map = StudentService._build_header_map(
            peeked[best_index]
        )

        remaining = itertools.chain(
            peeked[best_index + 1:],
            rows,
        )

        return header_map, remaining, best_index, peeked[best_index]

    @staticmethod
    def _parse_date(value):

        if value is None:
            return None

        if isinstance(value, datetime):

            return value.date()

        text = str(value).strip()

        if not text:
            return None

        for date_format in (
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y/%m/%d",
        ):

            try:

                return datetime.strptime(
                    text,
                    date_format,
                ).date()

            except ValueError:
                continue

        try:

            return datetime.fromisoformat(
                text,
            ).date()

        except ValueError:
            return None

    @staticmethod
    def _build_username(
        first_name,
        last_name,
        student_code,
    ):

        base = first_name or "student"

        if last_name:
            base = f"{first_name}_{last_name}"

        candidate = base.strip().replace(" ", "_").lower()

        if student_code:

            candidate = f"{candidate}_{student_code}"

        candidate = "".join(
            char for char in candidate
            if char.isalnum() or char == "_"
        )

        if not candidate:

            candidate = "student"

        unique_candidate = candidate

        counter = 1

        while User.objects.filter(
            username=unique_candidate,
        ).exists():

            unique_candidate = f"{candidate}_{counter}"

            counter += 1

        return unique_candidate

    @staticmethod
    def _build_student_code(
        institution,
        existing_code,
    ):

        if existing_code:

            code = existing_code.strip()

            if not Student.objects.filter(
                student_code=code,
            ).exists():

                return code

            return None

        prefix = "S"

        count = Student.objects.count()

        candidate = f"{prefix}{count + 1:06d}"

        while Student.objects.filter(
            student_code=candidate,
        ).exists():

            count += 1

            candidate = f"{prefix}{count + 1:06d}"

        return candidate

    @staticmethod
    def _parse_gender(value):
        """Normalize a gender value from the Excel file to male/female."""

        if value is None:
            return ""

        text = StudentService._normalize_header(
            str(value),
        )

        male_values = {
            "male",
            "m",
            "boy",
            "boys",
            "ذكر",
            "ولد",
            "بنين",
            "طلاب",
            "طالب",
        }

        female_values = {
            "female",
            "f",
            "girl",
            "girls",
            "انثى",
            "انثي",
            "بنت",
            "بنات",
            "طالبات",
            "طالبة",
            "طالبه",
        }

        if text in male_values:
            return Student.Gender.MALE

        if text in female_values:
            return Student.Gender.FEMALE

        return ""

    @staticmethod
    def _name_key(first_name, last_name):

        full_name = f"{first_name or ''} {last_name or ''}"

        return StudentService._normalize_header(full_name)

    @staticmethod
    def _split_full_name(full_name):

        parts = (full_name or "").split()

        if not parts:
            return "", ""

        first_name = parts[0]

        last_name = " ".join(parts[1:])

        return first_name, last_name

    @staticmethod
    @transaction.atomic
    def import_students(
        *,
        excel_file,
        institution,
        classroom,
        default_password="",
        default_gender="",
    ):

        if not excel_file.name.lower().endswith(
            (".xlsx", ".xls")
        ):

            raise ExcelImportError(
                _("Only Excel files (.xlsx or .xls) are allowed.")
            )

        try:

            workbook = load_workbook(
                excel_file,
                data_only=True,
                read_only=True,
            )

        except Exception:

            raise ExcelImportError(
                _(
                    "Could not read the Excel file. "
                    "Please make sure it is a valid Excel file."
                )
            )

        sheet = workbook.active

        rows = sheet.iter_rows(
            values_only=True,
        )

        header_map, data_rows, header_index, header_row = (
            StudentService._detect_header_row(rows)
        )

        if header_map is None:

            workbook.close()

            found_row = header_row[header_index] if (
                header_row and header_index >= 0
            ) else []

            found_columns = StudentService._format_found_columns(
                found_row,
            )

            raise ExcelImportError(
                _(
                    "Could not find the column headers in the file. "
                    "Found columns: {columns}. "
                    "Make sure the first row contains the column headers."
                ).format(
                    columns=found_columns,
                )
            )

        if (
            "first_name" not in header_map
            and "full_name" not in header_map
        ):

            workbook.close()

            found_columns = StudentService._format_found_columns(
                header_row,
            )

            raise ExcelImportError(
                _(
                    "Required column 'First Name' was not found in the file. "
                    "Found columns: {columns}. "
                    "Make sure the first row contains the column headers."
                ).format(
                    columns=found_columns,
                )
            )

        created_count = 0
        skipped_count = 0

        credentials = []

        errors = []

        seen_names = set()

        existing_keys = {
            StudentService._name_key(s.user.first_name, s.user.last_name)
            for s in Student.objects.filter(
                classroom=classroom,
            ).select_related("user")
        }

        for row_number, row in enumerate(
            data_rows,
            start=1,
        ):

            excel_row = header_index + 1 + row_number

            def get_value(field_name):

                index = header_map.get(field_name)

                if index is None or index >= len(row):
                    return None

                value = row[index]

                if value is None:
                    return None

                text = str(value).strip()

                return text if text else None

            first_name = get_value("first_name")
            last_name = get_value("last_name")

            if not first_name and not last_name:

                full_name = get_value("full_name")

                if full_name:

                    first_name, last_name = StudentService._split_full_name(
                        full_name,
                    )

            student_code = get_value("student_code")
            username = get_value("username")
            gender = StudentService._parse_gender(
                get_value("gender"),
            )
            date_of_birth = StudentService._parse_date(
                get_value("date_of_birth"),
            )
            parent_phone = get_value("parent_phone") or ""
            address = get_value("address") or ""

            if not first_name and not last_name:

                skipped_count += 1

                continue

            name_key = StudentService._name_key(
                first_name,
                last_name,
            )

            if name_key in seen_names or name_key in existing_keys:

                skipped_count += 1

                errors.append(
                    {
                        "row": excel_row,
                        "message": _(
                            "Student with the name '{name}' already exists."
                        ).format(
                            name=(
                                f"{first_name} {last_name or ''}"
                            ).strip(),
                        ),
                    }
                )

                continue

            student_code = StudentService._build_student_code(
                institution,
                student_code,
            )

            if student_code is None:

                skipped_count += 1

                errors.append(
                    {
                        "row": excel_row,
                        "message": _(
                            "Student code already exists for {name}."
                        ).format(
                            name=(
                                f"{first_name} {last_name or ''}"
                            ).strip(),
                        ),
                    }
                )

                continue

            if not username:

                username = StudentService._build_username(
                    first_name,
                    last_name,
                    student_code,
                )

            elif User.objects.filter(
                username=username,
            ).exists():

                skipped_count += 1

                errors.append(
                    {
                        "row": excel_row,
                        "message": _(
                            "Username '{username}' already exists for {name}."
                        ).format(
                            username=username,
                            name=(
                                f"{first_name} {last_name or ''}"
                            ).strip(),
                        ),
                    }
                )

                continue

            password = default_password or generate_strong_password()

            user = User.objects.create_user(
                username=username,
                password=password,
                first_name=first_name or "",
                last_name=last_name or "",
                role=User.Role.STUDENT,
                gender=gender or default_gender,
                must_change_password=True,
            )

            Student.objects.create(
                user=user,
                institution=institution,
                classroom=classroom,
                student_code=student_code,
                gender=gender or default_gender,
                date_of_birth=date_of_birth,
                address=address,
                parent_phone=parent_phone,
            )

            credentials.append(
                {
                    "code": user.login_code,
                    "username": username,
                    "name": (
                        f"{first_name or ''} {last_name or ''}"
                    ).strip(),
                    "password": password,
                }
            )

            seen_names.add(name_key)

            created_count += 1

        workbook.close()

        if created_count > 0:

            NotificationService.notify_institution_managers(
                institution,
                ntype=NotificationService.Type.STUDENT,
                title=_("Students imported"),
                message=_(
                    "{count} students were imported into {institution}."
                ).format(
                    count=created_count,
                    institution=institution.name,
                ),
                link="/students/",
            )

        return {
            "created_count": created_count,
            "skipped_count": skipped_count,
            "errors": errors,
            "credentials": credentials,
        }

    @staticmethod
    def _format_found_columns(header_row):

        if not header_row:
            return _("no columns found")

        columns = [
            str(cell).strip()
            for cell in header_row
            if cell is not None and str(cell).strip()
        ]

        if not columns:
            return _("no columns found")

        return ", ".join(columns)
